#!/usr/bin/env python3
"""
Build 2025-Modelos_720_y_721.xlsx from the 2024 Numbers-export template.
Values align with docs/private/finances/modelo_720_721_2025_updated_values.md.
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

TAX_YEAR = 2025
DEFAULT_TEMPLATE = Path(
    "/Users/markmhendrickson/Documents/data/imports/source_material/2024/taxes/2024-Modelos_720_y_721.xlsx"
)

# ECB / FX (Frankfurter + CoinGecko 2025-12-31)
RATE_USD_OCT1 = 0.85295
RATE_USD_DEC31 = 0.85106
RATE_BTC_EUR = 75269.84890335
RATE_STX_EUR = 0.21604444
RATE_UNI_EUR = 5.03852769

D_OCT1 = datetime(TAX_YEAR, 10, 1, 0, 0)
D_OCT2 = datetime(TAX_YEAR, 10, 2, 0, 0)
D_DEC31 = datetime(TAX_YEAR, 12, 31, 0, 0)

# --- Cantidades: fiat Q3 (rows 2–11) ---
FIAT_Q3 = [
    ("SEP", D_OCT1, 177_235.85, 151_172.32),
    ("Individual", D_OCT1, 513_539.03, 438_048.11),
    ("Roth", D_OCT1, 12_897.91, 11_001.27),
    ("Checking", D_OCT1, 11_279.21, 9_620.60),
    ("Tax", D_OCT2, 2.83, 2.41),
    ("Travel", D_OCT2, 6.26, 5.34),
    ("Big Expenses", D_OCT2, 113.08, 96.45),
    ("Charity", D_OCT2, 15.15, 12.92),
    ("Rainy day", D_OCT2, 9.17, 7.82),
    ("Lyft", D_OCT1, 1_120.18, 955.46),
]

# --- Cantidades: fiat Q4 (rows 13–22) ---
FIAT_Q4 = [
    ("SEP", 181_557.74, 154_517.49),
    ("Individual", 422_831.64, 359_855.38),
    ("Roth", 13_212.27, 11_244.26),
    ("Checking", 15_316.83, 13_035.81),
    ("Tax", 2.86, 2.43),
    ("Travel", 6.32, 5.38),
    ("Big Expenses", 114.05, 97.06),
    ("Charity", 15.27, 13.00),
    ("Rainy day", 9.26, 7.88),
    ("Lyft", 1_011.09, 860.47),
]

# --- Trimestre (rows 24–33) ---
FIAT_TRIM = [
    ("SEP", 179_396.80, 152_677.44),
    ("Individual", 468_185.34, 398_453.81),
    ("Roth", 13_055.09, 11_110.66),
    ("Checking", 13_298.02, 11_317.41),
    ("Tax", 2.85, 2.42),
    ("Travel", 6.29, 5.35),
    ("Big Expenses", 113.57, 96.65),
    ("Charity", 15.21, 12.94),
    ("Rainy day", 9.22, 7.84),
    ("Lyft", 1_065.64, 906.92),
]

EQUITY = [
    ("Hiro Systems PBC 1", 19_930),
    ("Hiro Systems PBC 2", 416),
    ("KITE Solutions, Inc. 1", 200_000),
    ("KITE Solutions, Inc. 2", 100_000),
    ("KITE Solutions, Inc. 3", 5_000),
    ("KITE Solutions, Inc. 4", 50_000),
    ("Leather Wallet LLC", 568_181),
    ("Nassau Machines LLC 1", 2_513),
    ("Nassau Machines LLC 2", 636),
]
DECL_720 = f"NO SE DECLARA MODELO 720-{TAX_YEAR}"

# (Bien label, units, sigla, eur, unit_eur_or_none) — match 2024 sheet labels (note trailing space on Bitcoin 1)
CRYPTO = [
    ("Stacks 1", 0, "STX", 0.0, RATE_STX_EUR),
    ("Stacks 2", 102_517.527851, "STX", 22_148.34, RATE_STX_EUR),
    ("Bitcoin ", 0.02989660, "BTC", 2_250.31, RATE_BTC_EUR),
    ("Bitcoin 2", 0.85830033, "BTC", 64_604.14, RATE_BTC_EUR),
    ("Bitcoin 3", 0, "BTC", 0.0, None),
    ("Bitcoin 4", 0, "BTC", 0.0, None),
    ("Bitcoin 5", 0.120763, "BTC", 9_089.81, RATE_BTC_EUR),
    ("Bitcoin 6", 0.01507159, "BTC", 1_134.44, RATE_BTC_EUR),
    ("UniSwap 1", 400, "UNI", 2_015.41, RATE_UNI_EUR),
    ("Coinbase BTC", 0.559, "BTC", 42_075.85, RATE_BTC_EUR),
    ("Coinbase STX", 0.003, "STX", 0.0, RATE_STX_EUR),
    ("Coinbase USDC", 261_175, "USDC", 222_275.60, None),
    # Kraken: verified `kraken_stocks_etfs_balances_2025-12-31.pdf` — no crypto; USD spot 0.0099 (negligible).
    ("Kraken BTC", 0, "BTC", 0.0, RATE_BTC_EUR),
    ("Kraken STX", 0, "STX", 0.0, RATE_STX_EUR),
    ("Kraken USDC", 0, "USDC", 0.0, None),
]


def write_fiat_block(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for i, (bien, fecha, usd, eur) in enumerate(FIAT_Q3):
        r = 2 + i
        ws.cell(r, 1, bien)
        ws.cell(r, 2, "Saldo")
        ws.cell(r, 3, fecha)
        ws.cell(r, 4, usd)
        ws.cell(r, 5, "USD")
        ws.cell(r, 6, eur)
        ws.cell(r, 7, None)

    for i, (bien, usd, eur) in enumerate(FIAT_Q4):
        r = 13 + i
        ws.cell(r, 1, bien)
        ws.cell(r, 2, "Saldo")
        ws.cell(r, 3, D_DEC31)
        ws.cell(r, 4, usd)
        ws.cell(r, 5, "USD")
        ws.cell(r, 6, eur)
        ws.cell(r, 7, "OK")

    for i, (bien, usd, eur) in enumerate(FIAT_TRIM):
        r = 24 + i
        ws.cell(r, 1, bien)
        ws.cell(r, 2, "Valor trimestre promedio")
        ws.cell(r, 3, D_DEC31)
        ws.cell(r, 4, usd)
        ws.cell(r, 5, "USD")
        ws.cell(r, 6, eur)
        ws.cell(r, 7, "OK")

    for i, (bien, units) in enumerate(EQUITY):
        r = 35 + i
        ws.cell(r, 1, bien)
        ws.cell(r, 2, "Valor")
        ws.cell(r, 3, D_DEC31)
        ws.cell(r, 4, units)
        ws.cell(r, 5, None)
        ws.cell(r, 6, "Ningún cambio")
        ws.cell(r, 7, DECL_720)

    crypto_start = 45
    total_units = 0.0
    total_eur = 0.0
    for i, (bien, units, sigla, eur, unit_eur) in enumerate(CRYPTO):
        r = crypto_start + i
        ws.cell(r, 1, bien)
        ws.cell(r, 2, "Valor")
        ws.cell(r, 3, D_DEC31)
        ws.cell(r, 4, units)
        ws.cell(r, 5, sigla)
        ws.cell(r, 6, eur)
        ws.cell(r, 7, unit_eur)
        total_units += float(units)
        total_eur += float(eur)

    last_crypto_row = crypto_start + len(CRYPTO) - 1
    totals_label_row = last_crypto_row + 2
    platform_row = totals_label_row + 2
    ratio_row = totals_label_row + 3

    # Copied template still has the prior-year footer (e.g. PLATAFORMA on row 60); clear
    # the variable block below the last crypto row before rewriting totals.
    clear_end = ratio_row + 3
    for r in range(last_crypto_row + 1, clear_end + 1):
        for c in range(1, 9):
            # openpyxl ignores ws.cell(r, c, None); must assign .value to clear formulas.
            ws.cell(r, c).value = None

    ws.cell(totals_label_row, 4, "unidades")
    ws.cell(totals_label_row, 6, "saldo")
    ws.cell(totals_label_row, 7, "valor unitario")

    ws.cell(platform_row, 1, "PLATAFORMA CRIPTOMONEDAS: ")
    ws.cell(platform_row, 4, total_units)
    ws.cell(platform_row, 6, total_eur)
    ratio = total_eur / total_units if total_units else None
    ws.cell(ratio_row, 6, ratio)


def update_export_summary(wb: openpyxl.workbook.workbook.Workbook, year: int) -> None:
    """Numbers export index still pointed at 2024 sheet names; align with output."""
    if "Export Summary" not in wb.sheetnames:
        return
    es = wb["Export Summary"]
    # Template rows 11–14: Numbers sheet name (B) and Excel worksheet name (D).
    es["B11"] = f"Cantidades {year}"
    es["D12"] = f"Cantidades {year}"
    es["B13"] = f"Conversiones {year}"
    es["D14"] = f"Conversiones {year}"


def write_conversiones(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.cell(1, 1, "Fecha")
    ws.cell(1, 2, "€ por unidad")
    ws.cell(1, 3, "Sigla")
    rows = [
        (D_OCT1, RATE_USD_OCT1, "USD"),
        (D_DEC31, RATE_USD_DEC31, "USD"),
        (D_DEC31, RATE_BTC_EUR, "BTC"),
        (D_DEC31, RATE_STX_EUR, "STX"),
        (D_DEC31, RATE_UNI_EUR, "UNI"),
    ]
    for i, (d, rate, sigla) in enumerate(rows, start=2):
        ws.cell(i, 1, d)
        ws.cell(i, 2, rate)
        ws.cell(i, 3, sigla)


def main() -> None:
    template = DEFAULT_TEMPLATE
    if not template.is_file():
        raise SystemExit(f"Template not found: {template}")

    out_desktop = (
        Path("/Users/markmhendrickson/Desktop") / f"{TAX_YEAR}-Modelos_720_y_721.xlsx"
    )
    out_source = (
        Path(
            "/Users/markmhendrickson/Documents/data/imports/source_material/2025/taxes"
        )
        / f"{TAX_YEAR}-Modelos_720_y_721.xlsx"
    )

    wb = load_workbook(template)

    for name in (f"Cantidades {TAX_YEAR}", f"Conversiones {TAX_YEAR}"):
        if name in wb.sheetnames:
            del wb[name]

    cant_src = wb["Cantidades 2024"]
    conv_src = wb["Conversiones 2024"]
    cant_new = wb.copy_worksheet(cant_src)
    cant_new.title = f"Cantidades {TAX_YEAR}"
    conv_new = wb.copy_worksheet(conv_src)
    conv_new.title = f"Conversiones {TAX_YEAR}"

    write_fiat_block(cant_new)
    write_conversiones(conv_new)

    update_export_summary(wb, TAX_YEAR)

    # Drop prior-year quantity sheets so the workbook is unambiguously 2025.
    wb.remove(wb["Cantidades 2024"])
    wb.remove(wb["Conversiones 2024"])

    wb.active = wb.sheetnames.index(cant_new.title)

    out_desktop.parent.mkdir(parents=True, exist_ok=True)
    out_source.parent.mkdir(parents=True, exist_ok=True)

    wb.save(out_desktop)
    shutil.copy2(out_desktop, out_source)

    print(f"Wrote {out_desktop}")
    print(f"Copied to {out_source}")


if __name__ == "__main__":
    main()
